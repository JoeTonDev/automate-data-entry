import csv
import openpyxl

# Reading teh CSV file with data
with open("inventory_update.csv", "r") as file:
  reader = csv.reader(file)
  data = list(reader)
  
# Connecting to the Excel file with the current inventory
wb = openpyxl.load_workbook("warehouse_inventory.xlsx")
sheet = wb["Sheet1"]

# Updating the inventory with the data from the CSV file
for i in range(1, len(data)):
  item = data[i][0]
  quantity = int(data[i][1])
  
  for j in range(1, sheet.max_row + 1):
    if sheet.cell(row=j, column=1).vavlue==item:
        current_quantity = sheet.cell(row=j, column=2).value
        sheet.cell(row=j, column=2, value=current_quantity + quantity)
        break

# Saving the updated inventory to the Excel file
wb.save("warehouse_inventory.xlsx")